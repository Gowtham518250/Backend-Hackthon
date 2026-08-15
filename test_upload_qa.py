#!/usr/bin/env python3
"""
Complete end-to-end test: Owner - Customer - Upload - Ask
"""

import requests
import time
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
import openpyxl

BASE_URL = "http://127.0.0.1:8001"

def create_test_files():
    """Create test PDF and Excel files."""
    # Create test PDF
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    c.setFont("Helvetica", 12)
    c.drawString(100, 750, "Sample Product Catalog")
    c.drawString(100, 730, "")
    c.drawString(100, 710, "Product 1: Laptop Computer")
    c.drawString(120, 690, "Price: $999.99")
    c.drawString(120, 670, "Specifications: 16GB RAM, 512GB SSD, Intel Core i7")
    c.drawString(100, 650, "")
    c.drawString(100, 630, "Product 2: Wireless Mouse")
    c.drawString(120, 610, "Price: $29.99")
    c.drawString(120, 590, "Specifications: Bluetooth, 12-month battery life")
    c.drawString(100, 570, "")
    c.drawString(100, 550, "Product 3: USB-C Hub")
    c.drawString(120, 530, "Price: $49.99")
    c.drawString(120, 510, "Specifications: 7 ports, supports 4K video output")
    c.showPage()
    c.save()
    
    with open("test_products.pdf", "wb") as f:
        f.write(pdf_buffer.getvalue())
    
    # Create test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    headers = ["Date", "Product", "Quantity", "Price", "Total"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    data = [
        ["2024-01-15", "Laptop Computer", 5, 999.99, 4999.95],
        ["2024-01-16", "Wireless Mouse", 12, 29.99, 359.88],
        ["2024-01-17", "USB-C Hub", 8, 49.99, 399.92],
        ["2024-01-18", "Laptop Computer", 3, 999.99, 2999.97],
        ["2024-01-19", "Keyboard", 15, 79.99, 1199.85],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("test_sales.xlsx")

def test_complete_flow():
    """Run the complete upload and Q&A flow."""
    
    # Create test files first
    print("Creating test PDF and Excel files...")
    create_test_files()
    print("✓ Test files created\n")
    print("COMPLETE E2E TEST: Owner - Customer - Upload - Ask")
    print("=" * 60)
    
    timestamp = int(time.time() * 1000)
    owner_email = f"owner{timestamp}@deepfind.com"
    employee_email = f"employee{timestamp}@deepfind.com"
    customer_email = f"customer{timestamp}@deepfind.com"
    
    # ===== STEP 1: Owner Registration =====
    print("\n[1] Owner Registration...")
    owner_resp = requests.post(
        f"{BASE_URL}/customer/register",
        data={
            "name": "Alice Owner",
            "email": owner_email,
            "password": "owner-password-123",
            "role": "owner"
        }
    )
    print(f"Status: {owner_resp.status_code}")
    owner_data = owner_resp.json()
    print(f"Owner ID: {owner_data.get('id')}")
    owner_id = owner_data.get('id')
    
    # ===== STEP 2: Owner Login =====
    print("\n[2] Owner Login...")
    login_resp = requests.post(
        f"{BASE_URL}/customer/login",
        data={
            "email": owner_email,
            "password": "owner-password-123"
        }
    )
    print(f"Status: {login_resp.status_code}")
    owner_token = login_resp.json().get('token')
    print(f"Token: {owner_token[:50]}...")
    owner_headers = {"Authorization": f"Bearer {owner_token}"}
    
    # ===== STEP 3: Create Customer/Organization =====
    print("\n[3] Owner Creates Customer Organization...")
    customer_org_resp = requests.post(
        f"{BASE_URL}/customer/create-customer",
        data={
            "name": "Acme Corporation",
            "organization_password": "org-secret-2025"
        },
        headers=owner_headers
    )
    print(f"Status: {customer_org_resp.status_code}")
    customer_org_data = customer_org_resp.json()
    print(f"Customer Response: {customer_org_data}")
    customer_id = customer_org_data.get('customer_id')
    
    # ===== STEP 4: Owner Creates Employee Account =====
    print("\n[4] Owner Creates Employee Account...")
    employee_resp = requests.post(
        f"{BASE_URL}/customer/create-user",
        data={
            "name": "Bob Employee",
            "email": employee_email,
            "password": "employee-pass-456",
            "role": "employee",
            "customer_id": customer_id
        },
        headers=owner_headers
    )
    print(f"Status: {employee_resp.status_code}")
    employee_data = employee_resp.json()
    print(f"Employee ID: {employee_data.get('id')}")
    employee_id = employee_data.get('id')
    
    # ===== STEP 5: Owner Creates Customer Account =====
    print("\n[5] Owner Creates Customer Account...")
    customer_user_resp = requests.post(
        f"{BASE_URL}/customer/create-user",
        data={
            "name": "Guest Customer",
            "email": customer_email,
            "password": "customer-pass-789",
            "role": "customer",
            "customer_id": customer_id
        },
        headers=owner_headers
    )
    print(f"Status: {customer_user_resp.status_code}")
    customer_user_data = customer_user_resp.json()
    print(f"Customer User ID: {customer_user_data.get('id')}")
    
    # ===== STEP 6: Employee Login =====
    print("\n[6] Employee Login...")
    employee_login_resp = requests.post(
        f"{BASE_URL}/customer/login",
        data={
            "email": employee_email,
            "password": "employee-pass-456"
        }
    )
    print(f"Status: {employee_login_resp.status_code}")
    employee_token = employee_login_resp.json().get('token')
    print(f"Token: {employee_token[:50]}...")
    employee_headers = {"Authorization": f"Bearer {employee_token}"}
    
    # ===== STEP 7: Employee Uploads PDF =====
    print("\n[7] Employee Uploads PDF File...")
    if os.path.exists("test_products.pdf"):
        with open("test_products.pdf", "rb") as f:
            files = {"file": ("test_products.pdf", f, "application/pdf")}
            pdf_upload_resp = requests.post(
                f"{BASE_URL}/customer/{customer_id}/upload",
                files=files,
                data={"organization_password": "org-secret-2025"},
                headers=employee_headers
            )
        print(f"Status: {pdf_upload_resp.status_code}")
        pdf_data = pdf_upload_resp.json()
        print(f"PDF Upload Response: {pdf_data}")
        pdf_file_id = pdf_data.get('file_id')
    else:
        print("❌ PDF file not found")
        pdf_file_id = None
    
    # ===== STEP 8: Employee Uploads Excel =====
    print("\n[8] Employee Uploads Excel File...")
    if os.path.exists("test_sales.xlsx"):
        with open("test_sales.xlsx", "rb") as f:
            files = {"file": ("test_sales.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            excel_upload_resp = requests.post(
                f"{BASE_URL}/customer/{customer_id}/upload",
                files=files,
                data={"organization_password": "org-secret-2025"},
                headers=employee_headers
            )
        print(f"Status: {excel_upload_resp.status_code}")
        excel_data = excel_upload_resp.json()
        print(f"Excel Upload Response: {excel_data}")
        excel_file_id = excel_data.get('file_id')
    else:
        print("❌ Excel file not found")
        excel_file_id = None
    
    # ===== STEP 9: Employee Lists Files =====
    print("\n[9] Employee Lists Customer Files...")
    files_list_resp = requests.get(
        f"{BASE_URL}/customer/{customer_id}/files?organization_password=org-secret-2025",
        headers=employee_headers
    )
    print(f"Status: {files_list_resp.status_code}")
    files_list = files_list_resp.json()
    print(f"Files: {files_list}")
    
    # ===== STEP 10: Employee Asks Question about Products =====
    print("\n[10] Employee Asks: 'What is the price of a Laptop Computer?'")
    question1_resp = requests.post(
        f"{BASE_URL}/customer/{customer_id}/ask",
        data={
            "question": "What is the price of a Laptop Computer?",
            "organization_password": "org-secret-2025"
        },
        headers=employee_headers
    )
    print(f"Status: {question1_resp.status_code}")
    answer1 = question1_resp.json()
    print(f"Answer: {answer1.get('answer')}")
    
    # ===== STEP 11: Employee Asks Question about Sales =====
    print("\n[11] Employee Asks: 'What was the total revenue from Laptop sales?'")
    question2_resp = requests.post(
        f"{BASE_URL}/customer/{customer_id}/ask",
        data={
            "question": "What was the total revenue from Laptop sales?",
            "organization_password": "org-secret-2025"
        },
        headers=employee_headers
    )
    print(f"Status: {question2_resp.status_code}")
    answer2 = question2_resp.json()
    print(f"Answer: {answer2.get('answer')}")
    
    # ===== STEP 12: Customer Login and Ask Question =====
    print("\n[12] Customer Login...")
    customer_login_resp = requests.post(
        f"{BASE_URL}/customer/login",
        data={
            "email": customer_email,
            "password": "customer-pass-789"
        }
    )
    print(f"Status: {customer_login_resp.status_code}")
    customer_token = customer_login_resp.json().get('token')
    customer_headers = {"Authorization": f"Bearer {customer_token}"}
    
    # ===== STEP 13: Customer Asks Question (Read-Only) =====
    print("\n[13] Customer Asks: 'List all products and their prices'")
    question3_resp = requests.post(
        f"{BASE_URL}/customer/{customer_id}/ask",
        data={
            "question": "List all products and their prices",
            "organization_password": "org-secret-2025"
        },
        headers=customer_headers
    )
    print(f"Status: {question3_resp.status_code}")
    answer3 = question3_resp.json()
    print(f"Answer: {answer3.get('answer')}")
    
    # ===== STEP 14: Customer Tries to Upload (Should Fail) =====
    print("\n[14] Customer Tries to Upload (Should Fail)...")
    if os.path.exists("test_products.pdf"):
        with open("test_products.pdf", "rb") as f:
            files = {"file": ("test_products.pdf", f, "application/pdf")}
            customer_upload_resp = requests.post(
                f"{BASE_URL}/customer/{customer_id}/upload",
                files=files,
                headers=customer_headers
            )
        print(f"Status: {customer_upload_resp.status_code}")
        print(f"Response: {customer_upload_resp.json()}")
        if customer_upload_resp.status_code == 403:
            print("✅ Correctly blocked customer from uploading")
        else:
            print("❌ Customer should not be able to upload!")
    
    print("\n" + "=" * 60)
    print("TEST COMPLETE")
    print("=" * 60)

if __name__ == "__main__":
    test_complete_flow()
